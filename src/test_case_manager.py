import os
import openpyxl


class TestCaseManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.reload()
        self.test_cases = {}
        self.current_test_case_name = None
        self.current_test_case_step_index = 0

    def load_test_cases_from_excel(self, excel_path):
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 1行目はヘッダーなので読み飛ばす
            if row[0]:  # Check if testcase name is not empty
                test_case_name = row[0]
                steps = [x for x in row[1:] if x is not None]
                self.test_cases[test_case_name] = steps

    def create_test_case_folders(self):
        for test_case_name in self.test_cases:
            os.makedirs(f"../data/{test_case_name}", exist_ok=True)

    def reload(self):
        self.wb = openpyxl.load_workbook(self.file_path)
        self.sheet = self.wb.active
        self.current_row = 2
        self.current_step = 1

    def get_next_test_case(self):
        if self.current_test_case_name is None:
            self.current_test_case_name = next(iter(self.test_cases))
        steps = self.test_cases.get(self.current_test_case_name)
        if steps is None:
            return None, None
        if self.current_test_case_step_index < len(steps):
            next_step = steps[self.current_test_case_step_index]
            self.current_test_case_step_index += 1
            return self.current_test_case_name, next_step
        else:
            self.current_test_case_name = None
            self.current_test_case_step_index = 0
            return self.get_next_test_case()
